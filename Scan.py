import cv2,os
from datetime import datetime, time, date
from openpyxl import Workbook, load_workbook
from PIL import Image

#Check sudah ada file Training atau belum
if not os.path.exists("Dataset/training.xml"):
  import Record

#Fungsi untuk menentukan waktu saat ini apakah dalam range waktu tertentu
def time_in_range(start, end, current):
    return start <= current <= end

'''
Define range waktu

ses_in_start = Start time untuk Check-In sesi
ses_in_end = End time untuk Check-In sesi

ses_out_start = Start time untuk Check-out sesi
ses_out_end = End time untuk Check-out sesi
'''

#Define range waktu masuk setiap sesi kelas
ses1_in_start = time(6,50,0)
ses1_in_end = time(7, 20, 0)

ses2_in_start = time(9,00,0)
ses2_in_end = time(9, 20, 0)

ses3_in_start = time(11,00,0)
ses3_in_end = time(11, 20, 0)

ses4_in_start = time(13,00,0)
ses4_in_end = time(13, 20, 0)

ses5_in_start = time(15,00,0)
ses5_in_end = time(15, 20, 0)

#Define range waktu keluar setiap sesi kelas
ses1_out_start = time(8,50,0)
ses1_out_end = time(8, 59, 59)

ses2_out_start = time(10,50,0)
ses2_out_end = time(10, 59, 59)

ses3_out_start = time(12,50,0)
ses3_out_end = time(12, 59, 59)

ses4_out_start = time(14,50,0)
ses4_out_end = time(14, 59, 59)

ses5_out_start = time(17,50,0)
ses5_out_end = time(18, 20, 0)

#Import Excel
wb = load_workbook('Attendance.xlsx')

#Import sheet excel
checkin = wb['Checkin']
checkout = wb['Checkout']

#Membuat list untuk record absensi
in_ses1=[]
out_ses1=[]

in_ses2=[]
out_ses2=[]

in_ses3=[]
out_ses3=[]

in_ses4=[]
out_ses4=[]

in_ses5=[]
out_ses5=[]

detect=0

def attendance_input(history, sheet, session):
    #Cek sebelumnya sudah check-in atau belum
    if id not in history:
        #Masukin ID ke list Check-in atau Check-Out supaya tidak absensi terus-menerus dalam satu sesi
        history.append(id)
        #Mengambil current time untuk di print ke excel
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        #Cari last row untuk formula vlookup
        LastRow = sheet.max_row +1
        #Append data ke excel
        sheet.append([id,'=VLOOKUP(A'+str(LastRow)+',\'Namelist\'!A:B,2,FALSE)',date.today(),session,current_time])
        
        #Update waktu last detect muka
        global last_detected
        last_detected=datetime.now()

        #Save file excel
        wb.save('Attendance.xlsx')
        
        #Update status detect
        global detect
        detect=1

        print('Absensi diterima!')

#Capture Webcam
cap = 0
video = cv2.VideoCapture(cap, cv2.CAP_DSHOW)

#Import Haar Cascade
FaceDetect = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

#Read .xml dari training data
Recog = cv2.face.LBPHFaceRecognizer_create()
Recog.read('Dataset/training.xml')
a=0

last_detected = datetime.now()

while True:
    a=a+1
    check, frame = video.read()
    #Ubah frame per frame ke Grayscale/Hitam putih
    grayscale = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    face = FaceDetect.detectMultiScale(grayscale,1.3,5)

    #Import current time buat nentuin sesi
    current = datetime.now().time()

    #Bikin rectangle dan ID di sekitar muka
    for (x1,y1,x2,y2) in face:
        cv2.rectangle(frame, (x1,y1), (x1+x2,y1+y2),(0,200,0),2)
        #Find ID dari muka
        id, conf = Recog.predict(grayscale[y1:y1+y2,x1:x1+x2])
        
        #Check-in Session 1
        if(time_in_range(ses1_in_start, ses1_in_end, current)):
            attendance_input(in_ses1, checkin, 'Session 1')
        #Check-out Session 1
        elif(time_in_range(ses1_out_start, ses1_out_end, current)):
            attendance_input(in_ses1, checkout,'Session 1')
        #Check-in Session 2
        elif(time_in_range(ses2_in_start, ses2_in_end, current)):
            attendance_input(in_ses2, checkin, 'Session 2')
        #Check-out Session 2
        elif(time_in_range(ses2_out_start, ses2_out_end, current)):
            attendance_input(in_ses2, checkout,'Session 2')
        #Check-in Session 3
        elif(time_in_range(ses3_in_start, ses3_in_end, current)):
            attendance_input(in_ses3, checkin, 'Session 3')
        #Check-out Session 3
        elif(time_in_range(ses3_out_start, ses3_out_end, current)):
            attendance_input(in_ses3, checkout,'Session 3')
        #Check-in Session 4
        elif(time_in_range(ses4_in_start, ses4_in_end, current)):
            attendance_input(in_ses4, checkin, 'Session 4')
        #Check-out Session 4
        elif(time_in_range(ses4_out_start, ses4_out_end, current)):
            attendance_input(in_ses4, checkout,'Session 4')
        #Check-in Session 5
        elif(time_in_range(ses5_in_start, ses5_in_end, current)):
            attendance_input(in_ses5, checkin, 'Session 5')
        #Check-out Session 5
        elif(time_in_range(ses5_out_start, ses5_out_end, current)):
            attendance_input(in_ses5, checkout,'Session 5')
        
        cv2.putText(frame,str(id),(x1+10,y1-10), cv2.FONT_HERSHEY_DUPLEX, 1, (0,200,0),2)
    if time_in_range(ses1_in_end, ses1_out_start, current) or time_in_range(ses2_in_end, ses2_out_start, current) or time_in_range(ses3_in_end, ses3_out_start, current) or time_in_range(ses4_in_end, ses4_out_start, current) or time_in_range(ses5_in_end, ses5_out_start, current):
        cv2.rectangle(frame, (215,5), (485,40),(0,0,255),-1)
        cv2.putText(frame, 'Session in Progress', (220, 30), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 2)
    if detect==1:
        if (datetime.now() - last_detected).total_seconds() < 3:
            cv2.rectangle(frame, (215,5), (450,40),(0,200,0),-1)
            cv2.putText(frame, 'Absensi diterima!', (220, 30), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 2)
        else:
            detect=0
    else:
        cv2.rectangle(frame, (15,410), (315,475),(0,0,0),-1)
        cv2.putText(frame, 'R : Register New Face', (20, 435), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 2)
        cv2.putText(frame, 'Space : Exit', (20, 465), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 2)
    #Show window buat webcam
    cv2.imshow("Attendance Cam", frame)
    #Waitkey 1 supaya video ga freeze
    key = cv2.waitKey(1)
    #Klik spasi jika ingin end program
    if key == ord(' '):
        break
    #Klik 'r' jika ingin daftar wajah baru
    elif key== ord('r'):
        cv2.destroyAllWindows()
        import Record
        Recog = cv2.face.LBPHFaceRecognizer_create()
        Recog.read('Dataset/training.xml')
    #Klik 'm' untuk manual Override untuk data Checkout Excel
    elif key==ord('m'):
        manual_id=int(input('Masukan ID: '))
        manual_session=input('Masukan sesi kelas: ')

        #Ngambil time buat di print ke excel
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        #Cari last row buat formula vlookup
        LastRow = checkout.max_row +1
        #Append data ke excel
        checkout.append([manual_id,'=VLOOKUP(A'+str(LastRow)+',\'Namelist\'!A:B,2,FALSE)',date.today(),'Session '+manual_session,current_time])

        #Save file excel
        wb.save('Attendance.xlsx')
        #Print notice manual override berhasil
        print('Input Data Check-Out berhasil')
cv2.destroyAllWindows()
video.release()